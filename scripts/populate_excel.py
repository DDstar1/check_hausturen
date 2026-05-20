"""
populate_excel.py — Fetch products from WooCommerce categories and pre-fill
the Excel masterlist with product IDs and prices.

Usage:
    python populate_excel.py path/to/masterlist.xlsx

Column sets are detected automatically from the Excel headers.
Any column ending in _id that has matching _base_price and _ab_price columns
is treated as a set. You will be prompted for the WooCommerce category slug
for each set found.
"""

import sys
from pathlib import Path

import requests
from requests.auth import HTTPBasicAuth
import openpyxl

from utils import WOO_STORES, find_id_columns, excel_value_to_slug, is_ab_price_showing, extract_ab_price, parse_price

# ── API helpers ───────────────────────────────────────────────────────────────

def _api_get(endpoint, store, params=None):
    auth = HTTPBasicAuth(store["consumer_key"], store["consumer_secret"])
    base = f"{store['url'].rstrip('/')}/wp-json/wc/v3"
    r = requests.get(f"{base}/{endpoint}", auth=auth, params=params or {}, timeout=20)
    r.raise_for_status()
    return r.json()


def _fetch_all(endpoint, store, extra=None):
    items, page = [], 1
    while True:
        batch = _api_get(endpoint, store, {"per_page": 100, "page": page, **(extra or {})})
        if not batch:
            break
        items.extend(batch)
        if len(batch) < 100:
            break
        page += 1
    return items


def _get_category_id(cat_slug, store):
    cats = _api_get("products/categories", store, {"slug": cat_slug})
    if not cats:
        raise ValueError(f"Category not found on site: {cat_slug!r}")
    return cats[0]["id"]


def _to_english(price_str) -> str:
    value = parse_price(price_str)
    if value is None:
        return ""
    return f"{value:.2f}"


def _prices(product):
    """Return (base_price_str, ab_price_str) in English format."""
    if is_ab_price_showing(product):
        return _to_english(product.get("price") or ""), _to_english(extract_ab_price(product))
    return _to_english(product.get("regular_price") or product.get("price") or ""), ""

# ── Store picker ──────────────────────────────────────────────────────────────

def _pick_store():
    names = list(WOO_STORES.keys())
    if len(names) == 1:
        return WOO_STORES[names[0]]
    print("Select store:")
    for i, n in enumerate(names, 1):
        print(f"  {i}. {n}")
    while True:
        try:
            idx = int(input("Enter number: ")) - 1
            if 0 <= idx < len(names):
                return WOO_STORES[names[idx]]
        except (ValueError, KeyboardInterrupt):
            pass

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python populate_excel.py path/to/masterlist.xlsx")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    # ── Open workbook and detect column sets ──────────────────────────────────
    print(f"Opening {xlsx_path.name} …")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
    headers_lower = [h.lower() for h in headers]

    def ci(name):
        try:
            return headers_lower.index(name.lower()) + 1
        except ValueError:
            return None

    name_ci = ci("PRODUCTS") or ci("Product") or ci("NAME") or ci("Produkt")
    if name_ci is None:
        print("ERROR: Could not find PRODUCTS / NAME column in row 1.")
        sys.exit(1)

    col_sets = find_id_columns(headers)
    valid_sets = [s for s in col_sets if s["base_price_col"] and s["ab_price_col"]]

    if not valid_sets:
        print("ERROR: No complete column sets found (need _id + _base_price + _ab_price).")
        sys.exit(1)

    print(f"\nFound {len(valid_sets)} column set(s):")
    for s in valid_sets:
        print(f"  {s['id_col']}  {s['base_price_col']}  {s['ab_price_col']}")

    # ── Ask for category slug per set ─────────────────────────────────────────
    print()
    for s in valid_sets:
        while True:
            slug = input(f"WooCommerce category slug for '{s['prefix']}': ").strip()
            if slug:
                s["cat_slug"] = slug
                break

    store = _pick_store()

    # ── Fetch products for each set ───────────────────────────────────────────
    cat_data = []
    for s in valid_sets:
        slug = s["cat_slug"]
        print(f"\nFetching  {slug} …", end=" ", flush=True)
        cat_id   = _get_category_id(slug, store)
        products = _fetch_all("products", store, {"category": cat_id, "status": "publish"})
        print(f"{len(products)} products")

        lookup = {}
        for p in products:
            lookup[p["slug"]] = p
            ns = excel_value_to_slug(p["name"])
            if ns:
                lookup.setdefault(ns, p)

        cat_data.append((s, cat_id, lookup, products))

    # ── Row iteration ─────────────────────────────────────────────────────────
    filled    = {s["cat_slug"]: 0 for s in valid_sets}
    not_found = []

    for row_num in range(2, ws.max_row + 1):
        raw = ws.cell(row=row_num, column=name_ci).value
        if not raw or str(raw).strip() in ("", "None", "-"):
            continue

        raw_str   = str(raw).strip()
        name_slug = excel_value_to_slug(raw_str)
        matched   = False

        for col_set, cat_id, lookup, products in cat_data:
            product = lookup.get(name_slug)
            if product is None:
                for p in products:
                    ps = p["slug"]
                    if name_slug in ps or ps.endswith(f"-{name_slug}") or ps.startswith(f"{name_slug}-"):
                        product = p
                        break
            if product is None:
                continue

            product_cat_ids = {c["id"] for c in product.get("categories", [])}
            if cat_id not in product_cat_ids:
                continue

            matched = True
            id_ci   = ci(col_set["id_col"])
            base_ci = ci(col_set["base_price_col"])
            ab_ci   = ci(col_set["ab_price_col"])
            base_price, ab_price = _prices(product)

            if id_ci:
                ws.cell(row=row_num, column=id_ci).value = product["id"]
            if base_ci:
                ws.cell(row=row_num, column=base_ci).value = base_price
            if ab_ci:
                ws.cell(row=row_num, column=ab_ci).value = ab_price
            filled[col_set["cat_slug"]] += 1

        if not matched:
            not_found.append(raw_str)

    # ── Save ──────────────────────────────────────────────────────────────────
    out_path = xlsx_path.with_stem(xlsx_path.stem + "_populated")
    wb.save(out_path)

    print(f"\nSaved -> {out_path}")
    for s in valid_sets:
        print(f"  {s['prefix']:20s}: {filled[s['cat_slug']]} rows filled")

    if not_found:
        print(f"\n{len(not_found)} product name(s) not matched to any category:")
        for n in not_found:
            print(f"  • {n}")


if __name__ == "__main__":
    main()
