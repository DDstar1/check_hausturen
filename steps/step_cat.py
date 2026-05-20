"""Step Cat — compare a WooCommerce category against the Excel masterlist."""

import os
import openpyxl

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFileDialog, QComboBox, QGroupBox, QLineEdit, QMessageBox,
    QRadioButton, QButtonGroup, QSizePolicy, QScrollArea,
)
from PySide6.QtCore import Qt

from utils import WOO_STORES, find_id_columns


class StepCat(QWidget):
    def __init__(self):
        super().__init__()
        self._wb = None
        self._col_sets = []
        self._table_map = {}
        self._build()

    def _build(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        outer.addWidget(scroll)

        inner = QWidget()
        scroll.setWidget(inner)
        layout = QVBoxLayout(inner)
        layout.setSpacing(16)
        layout.setContentsMargins(40, 28, 40, 28)

        title = QLabel("Find Missing Products")
        title.setObjectName("step_title")
        layout.addWidget(title)

        sub = QLabel(
            "Validate products in a WooCommerce category against your Excel masterlist."
        )
        sub.setObjectName("step_sub")
        sub.setWordWrap(True)
        layout.addWidget(sub)
        layout.addSpacing(8)

        # ── Store ────────────────────────────────────────────────────────────
        sg = QGroupBox("WooCommerce Store")
        sgl = QHBoxLayout(sg)
        self._store_combo = QComboBox()
        self._store_combo.addItems(list(WOO_STORES.keys()))
        self._store_url_lbl = QLabel()
        self._store_url_lbl.setStyleSheet("color: #6c7086; font-size: 11px;")
        self._store_combo.currentTextChanged.connect(self._on_store_changed)
        self._on_store_changed(self._store_combo.currentText())
        sgl.addWidget(self._store_combo)
        sgl.addWidget(self._store_url_lbl, stretch=1)
        layout.addWidget(sg)

        # ── Excel file ───────────────────────────────────────────────────────
        fg = QGroupBox("Excel File")
        fgl = QHBoxLayout(fg)
        self._file_label = QLabel("No file selected")
        self._file_label.setStyleSheet("color: #6c7086;")
        self._file_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        btn_browse = QPushButton("Browse…")
        btn_browse.clicked.connect(self._browse)
        fgl.addWidget(self._file_label)
        fgl.addWidget(btn_browse)
        layout.addWidget(fg)

        # ── Table ────────────────────────────────────────────────────────────
        tg = QGroupBox("Table")
        tgl = QVBoxLayout(tg)
        self._table_combo = QComboBox()
        self._table_combo.setEnabled(False)
        self._table_info_lbl = QLabel("")
        self._table_info_lbl.setStyleSheet("color: #6c7086; font-size: 11px;")
        self._table_combo.currentTextChanged.connect(self._on_table_changed)
        tgl.addWidget(self._table_combo)
        tgl.addWidget(self._table_info_lbl)
        layout.addWidget(tg)

        # ── Category slug ────────────────────────────────────────────────────
        cg = QGroupBox("WooCommerce Category Slug")
        cgl = QVBoxLayout(cg)
        self._slug_edit = QLineEdit()
        self._slug_edit.setPlaceholderText("e.g. rodenberg-premium-aluminium-haustueren")
        cgl.addWidget(self._slug_edit)
        layout.addWidget(cg)

        # ── Direction ────────────────────────────────────────────────────────
        dg = QGroupBox("Check Direction")
        dgl = QVBoxLayout(dg)
        self._btn_w2m = QRadioButton("Website → Masterlist  (find website products missing from Excel)")
        self._btn_m2w = QRadioButton("Masterlist → Website  (find Excel products missing from or wrong category on website)")
        self._btn_w2m.setChecked(True)
        self._dir_group = QButtonGroup(self)
        self._dir_group.addButton(self._btn_w2m)
        self._dir_group.addButton(self._btn_m2w)
        self._btn_m2w.toggled.connect(self._on_direction_changed)
        dgl.addWidget(self._btn_w2m)
        dgl.addWidget(self._btn_m2w)
        layout.addWidget(dg)

        # ── ID column set ─────────────────────────────────────────────────────
        self._col_group = QGroupBox("ID Column Set")
        col_l = QVBoxLayout(self._col_group)
        col_l.addWidget(QLabel("Which ID column should be used to match products?"))
        self._col_combo = QComboBox()
        col_l.addWidget(self._col_combo)
        layout.addWidget(self._col_group)

        layout.addStretch()

    # ── Slots ─────────────────────────────────────────────────────────────────

    def _on_store_changed(self, name):
        store = WOO_STORES.get(name, {})
        self._store_url_lbl.setText(store.get("url", ""))

    def _on_direction_changed(self, _):
        pass  # col group always visible

    def _browse(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel file", "", "Excel Files (*.xlsx *.xlsm)")
        if not path:
            return
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not open file:\n{e}")
            return

        table_map = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for tbl in ws.tables.values():
                table_map[tbl.name] = (sheet_name, tbl)

        if not table_map:
            QMessageBox.warning(self, "No tables found",
                "No Excel tables were found in this workbook.")
            return

        self._wb = wb
        self._table_map = table_map
        self._file_label.setText(os.path.basename(path))
        self._file_label.setStyleSheet("color: #cdd6f4;")
        self._table_combo.clear()
        self._table_combo.addItems(sorted(table_map.keys()))
        self._table_combo.setEnabled(True)

    def _on_table_changed(self, name):
        if name and name in self._table_map:
            sheet_name, tbl = self._table_map[name]
            self._table_info_lbl.setText(f"Sheet: {sheet_name}   Range: {tbl.ref}")
            self._refresh_col_sets(sheet_name, tbl.ref)
        else:
            self._table_info_lbl.setText("")

    def _refresh_col_sets(self, sheet_name, table_ref):
        if not self._wb:
            return
        ws = self._wb[sheet_name]
        headers = [str(c.value).strip() if c.value else ""
                   for c in ws[table_ref][0]]
        self._col_sets = [s for s in find_id_columns(headers)
                          if s["base_price_col"] and s["ab_price_col"]]
        self._col_combo.clear()
        for s in self._col_sets:
            self._col_combo.addItem(
                f"{s['id_col']}  ({s['base_price_col']}, {s['ab_price_col']})",
                userData=s,
            )

    # ── Public API ────────────────────────────────────────────────────────────

    def get_config(self) -> dict | None:
        store = WOO_STORES.get(self._store_combo.currentText())
        cat_slug = self._slug_edit.text().strip()
        table_name = self._table_combo.currentText()

        if not self._wb or not table_name or table_name not in self._table_map:
            QMessageBox.warning(self, "Incomplete", "Please select an Excel file and table.")
            return None
        if not cat_slug:
            QMessageBox.warning(self, "Incomplete", "Please enter a WooCommerce category slug.")
            return None

        sheet_name, tbl = self._table_map[table_name]
        mode = "masterlist_to_website" if self._btn_m2w.isChecked() else "website_to_masterlist"

        idx = self._col_combo.currentIndex()
        if idx < 0 or not self._col_sets:
            QMessageBox.warning(self, "No column set",
                "No ID column sets found in this Excel table.")
            return None
        col_set = self._col_combo.itemData(idx)

        return {
            "store":      store,
            "wb":         self._wb,
            "sheet_name": sheet_name,
            "table_ref":  tbl.ref,
            "cat_slug":   cat_slug,
            "col_set":    col_set,
            "mode":       mode,
        }
