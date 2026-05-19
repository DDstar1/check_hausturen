"""
populate_excel.py — Fetch products from WooCommerce categories and pre-fill
the Excel masterlist with product IDs and prices.

Usage:
    python populate_excel.py path/to/masterlist.xlsx

Columns filled:
    Aluminium_id  Aluminium_Base_Price  Aluminium_Ab_Price
    Kunstoff_id   Kunstoff_Base_Price   Kunstoff_Ab_Price

Matching: the PRODUCTS column value is slugified and compared against each
WooCommerce product's slug and slugified name.
"""

import sys
from pathlib import Path

import requests
from requests.auth import HTTPBasicAuth
import openpyxl

from utils import WOO_STORES, excel_value_to_slug, is_ab_price_showing, extract_ab_price, parse_price

# ── Target categories and the columns they map to ────────────────────────────

TARGETS = [
    {
        "cat_slug":    "rodenberg-premium-aluminium-haustuerfuellungen",
        "link_keyword": "aluminium",
        "id_col":      "Aluminium_id",
        "base_col":    "Aluminium_base_price",
        "ab_col":      "Aluminium_ab_price",
    },
    {
        "cat_slug":    "rodenberg-premium-kunststoff-haustuerfuellungen",
        "link_keyword": "kunststoff",
        "id_col":      "Kunstoff_id",
        "base_col":    "Kunstoff_base_price",
        "ab_col":      "Kunstoff_ab_price",
    },
]

# ── API helpers ───────────────────────────────────────────────────────────────

def _api_get(endpoint, store, params=None):
    auth = HTTPBasicAuth(store["consumer_key"], store["consumer_secret"])
    base = f"{store['url'].rstrip('/')}/wp-json/wc/v3"
    r = requests.get(f"{base}/{endpoint}", auth=auth, params=params or {}, timeout=20)
    r.raise_for_status()
    return r.json()


def _fetch_all(endpoint, store, extra=None):
    items, page = [], 1
    while True:
        batch = _api_get(endpoint, store, {"per_page": 100, "page": page, **(extra or {})})
        if not batch:
            break
        items.extend(batch)
        if len(batch) < 100:
            break
        page += 1
    return items


def _get_category_id(cat_slug, store):
    cats = _api_get("products/categories", store, {"slug": cat_slug})
    if not cats:
        raise ValueError(f"Category not found on site: {cat_slug!r}")
    return cats[0]["id"]


def _to_english(price_str) -> str:
    value = parse_price(price_str)
    if value is None:
        return ""
    return f"{value:.2f}"


def _prices(product):
    """Return (base_price_str, ab_price_str) in English format."""
    if is_ab_price_showing(product):
        return _to_english(product.get("price") or ""), _to_english(extract_ab_price(product))
    return _to_english(product.get("regular_price") or product.get("price") or ""), ""

# ── Store picker ──────────────────────────────────────────────────────────────

def _pick_store():
    names = list(WOO_STORES.keys())
    if len(names) == 1:
        return WOO_STORES[names[0]]
    print("Select store:")
    for i, n in enumerate(names, 1):
        print(f"  {i}. {n}")
    while True:
        try:
            idx = int(input("Enter number: ")) - 1
            if 0 <= idx < len(names):
                return WOO_STORES[names[idx]]
        except (ValueError, KeyboardInterrupt):
            pass

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python populate_excel.py path/to/masterlist.xlsx")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    store = _pick_store()

    # ── Fetch products for each category and build slug -> product lookup ──────
    cat_data = []  # list of (target_dict, lookup_dict, products_list)
    for target in TARGETS:
        slug = target["cat_slug"]
        print(f"\nFetching  {slug} …", end=" ", flush=True)
        cat_id   = _get_category_id(slug, store)
        products = _fetch_all("products", store, {"category": cat_id, "status": "publish"})
        print(f"{len(products)} products")

        lookup = {}
        for p in products:
            lookup[p["slug"]] = p                           # WC post slug
            ns = excel_value_to_slug(p["name"])
            if ns:
                lookup.setdefault(ns, p)                    # slugified name fallback

        # Debug: show a few WC slugs so mismatches are visible
        sample = list(lookup.keys())[:5]
        print(f"  Sample WC slugs: {sample}")

        cat_data.append((target, lookup, products))

    # ── Open workbook ─────────────────────────────────────────────────────────
    print(f"\nOpening {xlsx_path.name} …")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
    headers_lower = [h.lower() for h in headers]

    def ci(name):
        """1-based column index, case-insensitive, or None if not found."""
        try:
            return headers_lower.index(name.lower()) + 1
        except ValueError:
            return None

    # Find the product name column
    name_ci = ci("PRODUCTS") or ci("Product") or ci("NAME") or ci("Produkt")
    if name_ci is None:
        print("ERROR: Could not find PRODUCTS / NAME column in row 1.")
        sys.exit(1)

    # Warn about missing target columns
    for target, _lookup, _prods in cat_data:
        for key in ("id_col", "base_col", "ab_col"):
            col = target[key]
            if ci(col) is None:
                print(f"  WARNING: column '{col}' not found — rows will be skipped for it")

    # ── Row iteration ─────────────────────────────────────────────────────────
    filled    = {t["cat_slug"]: 0 for t in TARGETS}
    not_found = []

    # Show first few Excel name slugs for comparison
    sample_rows = []
    for row_num in range(2, min(7, ws.max_row + 1)):
        raw = ws.cell(row=row_num, column=name_ci).value
        if raw and str(raw).strip() not in ("", "None", "-"):
            sample_rows.append(f"{str(raw).strip()!r} -> {excel_value_to_slug(str(raw).strip())!r}")
    if sample_rows:
        print(f"\n  Sample Excel slugs: {sample_rows}")

    for row_num in range(2, ws.max_row + 1):
        raw = ws.cell(row=row_num, column=name_ci).value
        if not raw or str(raw).strip() in ("", "None", "-"):
            continue

        raw_str   = str(raw).strip()
        name_slug = excel_value_to_slug(raw_str)
        matched   = False

        for target, lookup, products in cat_data:
            # Exact slug match first, then substring fallback
            product = lookup.get(name_slug)
            if product is None:
                for p in products:
                    ps = p["slug"]
                    if name_slug in ps or ps.endswith(f"-{name_slug}") or ps.startswith(f"{name_slug}-"):
                        product = p
                        break
            if product is None:
                continue

            # Confirm the permalink contains the expected material keyword
            keyword = target.get("link_keyword", "")
            if keyword and keyword not in product.get("permalink", "").lower():
                print(f"  SKIP {product['id']} — permalink missing '{keyword}': {product.get('permalink', '')}")
                continue

            matched = True
            id_ci, base_ci, ab_ci = ci(target["id_col"]), ci(target["base_col"]), ci(target["ab_col"])
            base_price, ab_price  = _prices(product)

            if id_ci:
                ws.cell(row=row_num, column=id_ci).value = product["id"]
            if base_ci:
                ws.cell(row=row_num, column=base_ci).value = base_price
            if ab_ci:
                ws.cell(row=row_num, column=ab_ci).value = ab_price
            filled[target["cat_slug"]] += 1

        if not matched:
            not_found.append(raw_str)

    # ── Save ──────────────────────────────────────────────────────────────────
    out_path = xlsx_path.with_stem(xlsx_path.stem + "_populated")
    wb.save(out_path)

    print(f"\nSaved -> {out_path}")
    for cat_slug, count in filled.items():
        label = next(t["id_col"] for t in TARGETS if t["cat_slug"] == cat_slug)
        print(f"  {label.split('_')[0]:12s}: {count} rows filled")

    if not_found:
        print(f"\n{len(not_found)} product name(s) not matched to any category:")
        for n in not_found:
            print(f"  • {n}")


if __name__ == "__main__":
    main()
