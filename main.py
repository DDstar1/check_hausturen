from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from urllib.parse import urlparse
import time
import json
import csv
import re
import os
import openpyxl

START_URL = input("Enter product URL or category URL: ").strip().strip('"').strip("'")
_output_name = input("Enter output CSV file name (e.g. results): ").strip().strip('"').strip("'")
os.makedirs("results", exist_ok=True)
OUTPUT_FILE = os.path.join("results", os.path.splitext(_output_name)[0] + ".csv")
EXCEL_FILE = input("Enter Excel file path (e.g. Premium Masterlist A.xlsx): ").strip().strip('"').strip("'")
DOMAIN = "haustuerenshop.paultec.de"
CATEGORY_PREFIX = START_URL.rstrip("/")
_start_segments = [s for s in urlparse(START_URL).path.split("/") if s]
CATEGORY_SLUG = _start_segments[-1] if _start_segments else ""

SINGLE_PRODUCT_MODE = "/produkt/" in START_URL

visited = set()
product_links = {}  # {product_url: source_category_url}

# URL slug segment → Excel category name
URL_TO_CATEGORY = {
    "modern":     "MODERN",
    "exklusiv":   "EXCLUSIV",
    "klassisch":  "KLASSICH",
    "lignum":     "LIGNUM",
    "loftdesign": "LOFTDESIGN",
    "ganzglas":   "GLAS & ELEGANZ",
}

URL_TO_MATERIAL = {
    "aluminium":  "aluminium",
    "ganzglas":   "ganzglas",
    "kunststoff": "kunststoff",
}

def category_from_url(url):
    for slug, cat in URL_TO_CATEGORY.items():
        if slug in url:
            return cat
    return None

def material_from_url(url):
    for slug, mat in URL_TO_MATERIAL.items():
        if slug in url:
            return mat
    return None

ATTR_COLUMN_MAP = {
    "FÜLLUNG":           "rodenberg-fuellung",
    "AUSFÜHRUNG":        "rodenberg-ausfuhrung",
    "BREITE":            "rodenberg-breite",
    "HÖHE":              "rodenberg-hoehe",
    "ELEMENT":           "rodenberg-element",
    "VERGLASUNG":        "rodenberg-verglasung",
    "LISENEN":           "rodenberg-lisenen",
    "FARBE":             "rodenberg-farbe",
    "Seitenteil RC 2/3": "rodenberg-seitenteil-breite",
}


SLUG_ALIASES = {
    "turblatt": "tuerfuellung",
    "aufsatzfuellung-seitenteil": "aufsatzfuellung",
}


def excel_value_to_slug(value):
    if not value or str(value).strip() in ("", "None", "-", "Nil"):
        return None
    value = str(value).strip()
    if re.match(r"Group\s+\d+", value, re.IGNORECASE):
        return "premium-group-" + re.search(r"\d+", value).group()
    if re.match(r"Premium\s+\d+", value, re.IGNORECASE):
        return "premium-" + re.search(r"\d+", value).group()
    slug = value.lower()
    for old, new in [("ü", "ue"), ("ö", "oe"), ("ä", "ae"), ("ß", "ss"), (" & ", "-"), ("&", "")]:
        slug = slug.replace(old, new)
    slug = re.sub(r"[^\w-]", "-", slug)
    slug = re.sub(r"-+", "-", slug).strip("-")
    for alias_from, alias_to in SLUG_ALIASES.items():
        slug = slug.replace(alias_from, alias_to)
    return slug


# --- Load Excel master list ---
print(f"\nLoading Excel: {EXCEL_FILE}")
excel_products = {}

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb.active
header_row = [str(c.value).strip() if c.value else "" for c in ws[1]]
print(f"  Headers found: {[h for h in header_row if h]}")

for row in ws.iter_rows(min_row=2, values_only=True):
    product_name = str(row[0]).strip() if row[0] else ""
    if not product_name or product_name == "None":
        continue
    category = str(row[1]).strip() if row[1] else ""

    m = re.search(r"Typ\s+(\d+)\s*-\s*(\d+)", product_name, re.IGNORECASE)
    code = f"Typ {m.group(1)}-{m.group(2)}" if m else product_name

    expected_attr_values = {}
    for col_name, attr_slug in ATTR_COLUMN_MAP.items():
        if col_name in header_row:
            col_idx = header_row.index(col_name)
            val = row[col_idx] if col_idx < len(row) else None
            slug = excel_value_to_slug(val)
            if slug:
                expected_attr_values[attr_slug] = slug

    entry = {
        "name": product_name,
        "category": category,
        "expected_attr_values": expected_attr_values,
    }
    # same code can appear in multiple categories — store all as a list
    excel_products.setdefault(code, []).append(entry)

print(f"  Loaded {len(excel_products)} products from Excel.\n")

# --- Setup Selenium ---
print("Starting Chrome...")
options = Options()
options.add_argument("--disable-gpu")
options.add_argument("--headless")
driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 10)
print("Chrome ready.\n")

RETRY_DURATION = 600
RETRY_INTERVAL = 5


def is_valid_link(url):
    parsed = urlparse(url)
    return DOMAIN in parsed.netloc


def load_with_retry(url):
    deadline = time.time() + RETRY_DURATION
    while True:
        try:
            driver.get(url)
            return True
        except Exception as e:
            if time.time() >= deadline:
                print(f"  [ERROR] Network error on {url} — gave up after 10 minutes: {e}")
                return False
            print(f"  [RETRY] Network error — retrying in {RETRY_INTERVAL}s: {e}")
            time.sleep(RETRY_INTERVAL)


def crawl(url, source_category_url=None):
    if url in visited:
        return

    print(f"[CRAWL] {url}")
    visited.add(url)

    try:
        if not load_with_retry(url):
            return

        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))

        # wait for product grid or any product link to appear
        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "ul.products li.product, a[href*='/produkt/']")))
            print("  Product elements loaded.")
        except Exception:
            print("  WARNING: product grid not detected, continuing anyway.")

        hrefs = [el.get_attribute("href") for el in driver.find_elements(By.TAG_NAME, "a")]

        all_hrefs = [h for h in hrefs if h]
        print(f"  Total links on page: {len(all_hrefs)}")
        produkt_hrefs = [h for h in all_hrefs if "/produkt/" in h and is_valid_link(h)]
        print(f"  /produkt/ links found: {len(produkt_hrefs)}")

        for href in all_hrefs:
            if "#" in href:
                continue
            parsed_href = urlparse(href)
            if parsed_href.query:
                continue
            if "/produkt/" in href and is_valid_link(href):
                if href not in product_links:
                    product_links[href] = source_category_url or url
                    print(f"  [FOUND] Product #{len(product_links)}: {href} (from {product_links[href]})")
            elif "/produkte/" in href and CATEGORY_SLUG in href and is_valid_link(href):
                if href not in visited:
                    crawl(href, source_category_url=href)

    except Exception as e:
        print(f"  [ERROR] crawl failed on {url}: {e}")


def extract_and_validate(url, source_category_url, index, total):
    print(f"\n{'='*60}")
    print(f"[{index}/{total}] Loading: {url}")
    try:
        if not load_with_retry(url):
            print("  [ERROR] Could not load page, skipping.")
            return

        print("  Waiting for product form to appear...")
        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".product-summary.tc-init > form")))
            print("  Product form found.")
        except Exception:
            print("  WARNING: product form selector timed out, continuing anyway.")

        # --- Title ---
        title = ""
        try:
            title = driver.find_element(By.CSS_SELECTOR, "h1.product-title").text.strip()
            print(f"  Title: {title}")
        except Exception:
            print("  WARNING: could not find h1.product-title")

        # --- Button disabled ---
        button_disabled = None
        try:
            button = driver.find_element(By.CSS_SELECTOR, "button.single_add_to_cart_button")
            classes = button.get_attribute("class") or ""
            aria_disabled = button.get_attribute("aria-disabled") or ""
            button_disabled = "disabled" in classes or aria_disabled == "true"
            print(f"  Button classes: '{classes}' | aria-disabled: '{aria_disabled}' => disabled={button_disabled}")
        except Exception:
            print("  WARNING: buy button not found")

        # --- AB price ---
        ab_price_showing = False
        ab_price_value = ""
        try:
            price_p = driver.find_elements(By.CSS_SELECTOR, "div.price-wrapper > p")
            if price_p:
                from_span = price_p[0].find_elements(By.CSS_SELECTOR, "span.from")
                ab_price_showing = len(from_span) > 0
                if ab_price_showing:
                    try:
                        ab_price_value = price_p[0].find_element(
                            By.CSS_SELECTOR, "span.woocommerce-Price-amount bdi"
                        ).text.strip()
                    except Exception:
                        ab_price_value = price_p[0].text.strip()
                    print(f"  AB price: YES — {ab_price_value}")
                else:
                    print("  AB price: NO (no span.from in price-wrapper > p)")
            else:
                print("  AB price: NO (div.price-wrapper > p not found)")
        except Exception as e:
            print(f"  WARNING: error checking AB price: {e}")

        # --- Product form (li.tm-extra-product-options-field) ---
        has_form = False
        product_form_added = False
        try:
            lis = driver.find_elements(By.CSS_SELECTOR, "li.tm-extra-product-options-field")
            has_form = len(lis) > 0
            print(f"  TM extra options li count: {len(lis)}")
            if has_form:
                empty_count = 0
                for i, li in enumerate(lis):
                    height = driver.execute_script("return arguments[0].getBoundingClientRect().height", li)
                    li_id = li.get_attribute("id") or f"li[{i}]"
                    is_empty = height < 10
                    print(f"    {li_id}: height={height}px {'(EMPTY — form not added)' if is_empty else '(has content)'}")
                    if is_empty:
                        empty_count += 1
                product_form_added = empty_count == 0
                print(f"  => has_form={has_form}, product_form_added={product_form_added} ({empty_count} collapsed li(s) under 10px)")
            else:
                print("  => No li.tm-extra-product-options-field found at all")
        except Exception as e:
            print(f"  WARNING: error checking form li elements: {e}")

        # --- Variation attributes ---
        website_attr_values = {}
        try:
            form = driver.find_element(By.CSS_SELECTOR, "form.variations_form")
            variations_raw = form.get_attribute("data-product_variations")
            if variations_raw:
                for v in json.loads(variations_raw):
                    for key, val in v.get("attributes", {}).items():
                        k = key.lower()
                        website_attr_values.setdefault(k, set()).add(val.lower())
                print(f"  Website attributes found ({len(website_attr_values)}):")
                for k, vals in website_attr_values.items():
                    print(f"    {k}: {vals}")
            else:
                print("  WARNING: data-product_variations is empty")
        except Exception as e:
            print(f"  WARNING: could not extract variation attributes: {e}")

        # --- Match to Excel ---
        cm = re.search(r"Typ\s+(\d+)\s*-\s*(\d+)", title, re.IGNORECASE)
        excel_code = f"Typ {cm.group(1)}-{cm.group(2)}" if cm else None
        print(f"  Excel lookup code: {excel_code!r}")
        excel_entries = excel_products.get(excel_code) if excel_code else None

        if excel_entries:
            # narrow to the entry whose category matches the crawled category URL
            url_category = category_from_url(source_category_url or "")
            print(f"  URL-derived category: {url_category!r}")
            if url_category and len(excel_entries) > 1:
                matched = [e for e in excel_entries if e["category"].upper() == url_category.upper()]
                excel_entries = matched if matched else excel_entries
            print(f"  Excel entries to check: {len(excel_entries)}")
            for excel_entry in excel_entries:
                product_name_out = excel_entry["name"]
                category_out = excel_entry["category"]
                expected_attr_values = excel_entry["expected_attr_values"]
                print(f"  Checking entry: {product_name_out} | category: {category_out}")
                print(f"  Expected attr values: {expected_attr_values}")

                # build the full set of checks: Excel attrs + material from URL
                checks = dict(expected_attr_values)
                expected_material = material_from_url(source_category_url or "")
                if expected_material:
                    checks["material"] = expected_material
                    print(f"  Material check from URL: '{expected_material}'")

                if checks:
                    variables_match = True
                    for attr_slug, expected_slug in checks.items():
                        matching_key = next((k for k in website_attr_values if attr_slug in k), None)
                        website_vals = website_attr_values.get(matching_key, set()) if matching_key else set()
                        match_result = expected_slug in website_vals
                        print(f"    {attr_slug}: expected '{expected_slug}' | website key '{matching_key}' | values {website_vals} => {'OK' if match_result else 'MISMATCH'}")
                        if not match_result:
                            variables_match = False
                else:
                    variables_match = True
                    print("  No expected attrs in Excel row — skipping variable check")

                print(f"\n  RESULT: disabled={button_disabled} | ab={ab_price_showing} ({ab_price_value}) | has_form={has_form} | form_added={product_form_added} | vars={variables_match}")
                results.append([
                    product_name_out,
                    category_out,
                    button_disabled,
                    ab_price_showing,
                    ab_price_value,
                    has_form,
                    product_form_added,
                    variables_match,
                    url,
                ])
        else:
            print(f"  WARNING: No Excel entry found for code '{excel_code}'")
            print(f"\n  RESULT: disabled={button_disabled} | ab={ab_price_showing} ({ab_price_value}) | has_form={has_form} | form_added={product_form_added} | vars=NO MATCH IN EXCEL")
            results.append([
                title,
                "",
                button_disabled,
                ab_price_showing,
                ab_price_value,
                has_form,
                product_form_added,
                "NO MATCH IN EXCEL",
                url,
            ])

    except Exception as e:
        print(f"  [ERROR] Failed to validate {url}: {e}")
        results.append(["ERROR", "", "", "", "", "", "", "", url])


# --- Run ---
if SINGLE_PRODUCT_MODE:
    print(f"[SINGLE PRODUCT MODE] Skipping crawl, validating directly: {START_URL}\n")
    product_links[START_URL] = START_URL
else:
    print(f"[CATEGORY MODE] Starting crawl... (boundary slug: '{CATEGORY_SLUG}')\n")
    crawl(START_URL)
    print(f"\nFound {len(product_links)} product links\n")

results = []
total = len(product_links)
for index, (link, source_cat) in enumerate(product_links.items(), start=1):
    extract_and_validate(link, source_cat, index, total)

print(f"\nAll products checked. Writing CSV to {OUTPUT_FILE}...")
with open(OUTPUT_FILE, "w", newline="", encoding="utf-8") as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow([
        "Product Name",
        "Category",
        "Button Disabled",
        "AB Price Showing",
        "AB Price Value",
        "Has Form (li exists)",
        "Product Form Added",
        "Variables Match",
        "URL",
    ])
    writer.writerows(results)

print(f"\nDone. Results saved to {OUTPUT_FILE}")
driver.quit()
